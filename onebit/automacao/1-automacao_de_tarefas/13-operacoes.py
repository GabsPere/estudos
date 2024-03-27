# Realizando operações na planilha através do Python.

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

wb = load_workbook(filename='files/gastos.xlsx')
planilha = wb['Gastos']

valor_total = 0
# 1 -> Soma de valores
for i in range(2,7):
    valor = float(planilha['B%s' %i].value)
    valor_total += valor
print(f'Valor total de gastos R${valor_total:.2f}')
planilha["B7"] = valor_total

# 2 -> Mesclar Células
planilha['A8'] = 'Teste'
planilha.merge_cells('A8:B8')
# planilha.unmerge_cells('A8:B8') # -> Para desfazer

# 3 -> Inserindo imagem
img = Image('files/bb_preco.png')
planilha.add_image(img,'B9') # -> Só abre no proprio excel, não dá ver pelo viwer no vscode

# 4 -> Deletando Células
planilha.delete_rows(1)
planilha.delete_cols(3)

wb.save(filename='files/gastos2.xlsx') # -> Criando nova planilha com as linhas excluidas.