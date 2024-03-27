# Juntando planilhas em um único arquivo.

from openpyxl import load_workbook, Workbook

lista_arquivos = ['gastos','gastos2','teste'] # -> Planilhas que serão reunidas

# 1 -> Criando nova planilha

wb = Workbook()
nome_arquivo = 'files/juncao.xlsx'

for nome in lista_arquivos:
    arquivo = load_workbook(filename='files/%s.xlsx' %nome)
    sheet = arquivo[nome]
    max_linhas = sheet.max_row
    max_colunas = sheet.max_column
    ws = wb.create_sheet(title=nome)

    # 2 -> Iterando valores da planilha
    for linhas in range(1,max_linhas+1):
        for colunas in range(1,max_colunas+1):
            data = sheet.cell(row=linhas,column=colunas)
            ws.cell(row=linhas,column=colunas).value = data.value
wb.remove(wb['Sheet'])
wb.save(nome_arquivo)